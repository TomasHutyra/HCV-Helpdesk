import io
import os

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models as db_models
from django.http import Http404, HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.translation import gettext as _
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView

from apps.accounts.models import UserRole
from .forms import (
    TicketCreateForm, TicketUpdateForm, AssignResolverForm, AssignSalesForm,
    ResolveForm, RejectForm, ChangeTypeForm, CommentForm, TimeLogForm,
    TicketFilterForm, AttachmentUploadForm, WorkCategoryAdminForm,
)
from .models import Ticket, Comment, TimeLog, TicketAttachment, TicketChange, WorkCategory, ALLOWED_EXTENSIONS, MAX_UPLOAD_SIZE


def _log_change(ticket, user, field, old_value='', new_value=''):
    TicketChange.objects.create(
        ticket=ticket, user=user, field=field,
        old_value=str(old_value), new_value=str(new_value),
    )


def _status_label(value):
    return str(dict(Ticket.STATUS_CHOICES).get(value, value))


def _type_label(value):
    return str(dict(Ticket.TYPE_CHOICES).get(value, value))


def _priority_label(value):
    return str(dict(Ticket.PRIORITY_CHOICES).get(value, value))


def _validate_upload(f):
    """Vrátí chybovou zprávu nebo None pokud je soubor v pořádku."""
    ext = os.path.splitext(f.name)[1].lower().lstrip('.')
    if ext not in ALLOWED_EXTENSIONS:
        return _('Nepodporovaný typ souboru .%(ext)s.') % {'ext': ext}
    if f.size > MAX_UPLOAD_SIZE:
        return _('Soubor %(name)s je příliš velký (max 5 MB).') % {'name': f.name}
    return None


def _can_add_attachment(user, ticket):
    """Smí uživatel přidávat přílohy k tiketu?"""
    if user.has_role(UserRole.ADMIN):
        return True
    if user.has_role(UserRole.MANAGER):
        return user.can_see_ticket_as_manager(ticket)
    if user.has_role(UserRole.RESOLVER):
        return ticket.resolver == user
    if user.has_role(UserRole.SALES):
        return ticket.sales == user
    if user.has_role(UserRole.REQUESTER):
        return ticket.requester == user
    return False


def _can_delete_attachment(user, attachment):
    """Smí uživatel smazat přílohu? (nahrávatel nebo správce/admin)"""
    if user.has_role(UserRole.ADMIN):
        return True
    if user.has_role(UserRole.MANAGER):
        return user.can_see_ticket_as_manager(attachment.ticket)
    return attachment.uploaded_by == user


def _get_adjacent_tickets(user, ticket):
    """Vrátí (prev_pk, next_pk) — sousední tikety ve stejném pořadí jako seznam (dle PK)."""
    qs = Ticket.objects.all()
    if user.has_role(UserRole.ADMIN):
        pass  # Admin vidí vše
    elif user.has_role(UserRole.MANAGER):
        q = user.get_ticket_visibility_q()
        if q is not None:
            qs = qs.filter(q)
    elif user.has_role(UserRole.RESOLVER):
        area_q = user.get_resolver_new_tickets_q()
        new_q = db_models.Q(status=Ticket.STATUS_NEW) if area_q is None else db_models.Q(status=Ticket.STATUS_NEW) & area_q
        qs = qs.filter(db_models.Q(resolver=user) | new_q)
    elif user.has_role(UserRole.SALES):
        qs = qs.filter(sales=user)
    elif user.has_role(UserRole.REQUESTER):
        qs = qs.filter(requester=user)
    else:
        return None, None
    prev_pk = qs.filter(pk__lt=ticket.pk).order_by('-pk').values_list('pk', flat=True).first()
    next_pk = qs.filter(pk__gt=ticket.pk).order_by('pk').values_list('pk', flat=True).first()
    return prev_pk, next_pk


def _can_edit_ticket(user, ticket):
    """Může uživatel editovat tiket (tj. není v uzamčeném stavu)?"""
    if ticket.is_locked:
        return False
    if user.has_role(UserRole.ADMIN):
        return True
    if user.has_role(UserRole.MANAGER):
        return user.can_see_ticket_as_manager(ticket)
    if user.has_role(UserRole.RESOLVER) and ticket.resolver == user:
        return True
    return False


def _manager_has_ticket_access(user, ticket):
    """Vrátí True, pokud je uživatel správce (nebo admin) a smí přistoupit k tiketu."""
    if user.has_role(UserRole.ADMIN):
        return True
    if user.has_role(UserRole.MANAGER):
        return user.can_see_ticket_as_manager(ticket)
    return False


def _apply_ticket_filters(qs, get_params, user):
    """Aplikuje filtry z GET parametrů na queryset tiketů."""
    form = TicketFilterForm(get_params, user=user)
    if form.is_valid():
        if form.cleaned_data.get('status'):
            qs = qs.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('type'):
            qs = qs.filter(type=form.cleaned_data['type'])
        if form.cleaned_data.get('area'):
            qs = qs.filter(area=form.cleaned_data['area'])
        if form.cleaned_data.get('priority'):
            qs = qs.filter(priority=form.cleaned_data['priority'])
        if form.cleaned_data.get('company'):
            qs = qs.filter(company=form.cleaned_data['company'])
        if form.cleaned_data.get('requester'):
            qs = qs.filter(requester=form.cleaned_data['requester'])
        if form.cleaned_data.get('resolver'):
            qs = qs.filter(resolver=form.cleaned_data['resolver'])
        if form.cleaned_data.get('search'):
            q = form.cleaned_data['search']
            qs = qs.filter(
                db_models.Q(title__icontains=q) | db_models.Q(description__icontains=q)
            )
        if form.cleaned_data.get('date_from'):
            qs = qs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            qs = qs.filter(created_at__date__lte=form.cleaned_data['date_to'])
    return qs


class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'tickets/ticket_list.html'
    context_object_name = 'tickets'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        qs = Ticket.objects.select_related('requester', 'resolver', 'sales', 'company', 'area')

        if user.has_role(UserRole.ADMIN):
            pass
        elif user.has_role(UserRole.MANAGER):
            q = user.get_ticket_visibility_q()
            if q is not None:
                qs = qs.filter(q)
        elif user.has_role(UserRole.RESOLVER):
            area_q = user.get_resolver_new_tickets_q()
            new_q = db_models.Q(status=Ticket.STATUS_NEW) if area_q is None else db_models.Q(status=Ticket.STATUS_NEW) & area_q
            qs = qs.filter(db_models.Q(resolver=user) | new_q)
        elif user.has_role(UserRole.SALES):
            qs = qs.filter(sales=user)
        elif user.has_role(UserRole.REQUESTER):
            qs = qs.filter(requester=user)
        else:
            qs = qs.none()

        qs = _apply_ticket_filters(qs, self.request.GET, user)

        if user.has_role(UserRole.MANAGER, UserRole.RESOLVER, UserRole.SALES, UserRole.ADMIN):
            qs = qs.annotate(hours_sum=db_models.Sum('time_logs__hours'))

        return qs.order_by('-created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = TicketFilterForm(self.request.GET, user=self.request.user)
        ctx['show_hours'] = self.request.user.has_role(
            UserRole.MANAGER, UserRole.RESOLVER, UserRole.SALES, UserRole.ADMIN
        )
        return ctx


class TicketExportView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.utils.timezone import localtime

        user = request.user
        qs = Ticket.objects.select_related('requester', 'resolver', 'sales', 'company', 'area')

        if user.has_role(UserRole.ADMIN):
            pass
        elif user.has_role(UserRole.MANAGER):
            q = user.get_ticket_visibility_q()
            if q is not None:
                qs = qs.filter(q)
        elif user.has_role(UserRole.RESOLVER):
            area_q = user.get_resolver_new_tickets_q()
            new_q = db_models.Q(status=Ticket.STATUS_NEW) if area_q is None else db_models.Q(status=Ticket.STATUS_NEW) & area_q
            qs = qs.filter(db_models.Q(resolver=user) | new_q)
        elif user.has_role(UserRole.SALES):
            qs = qs.filter(sales=user)
        elif user.has_role(UserRole.REQUESTER):
            qs = qs.filter(requester=user)
        else:
            qs = qs.none()

        qs = _apply_ticket_filters(qs, request.GET, user)

        show_hours = user.has_role(UserRole.MANAGER, UserRole.RESOLVER, UserRole.SALES, UserRole.ADMIN)
        show_resolver = True

        if show_hours:
            qs = qs.annotate(hours_sum=db_models.Sum('time_logs__hours'))

        qs = qs.order_by('-created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Tikety')

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1A56DB')
        header_align = Alignment(horizontal='center', vertical='center')

        headers = ['#', _('Název'), _('Typ'), _('Stav'), _('Priorita'), _('Oblast'), _('Firma'), _('Žadatel')]
        if show_resolver:
            headers.append(_('Řešitel'))
        if show_hours:
            headers.append(_('Hodiny'))
        headers.append(_('Vytvořeno'))

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        type_labels = dict(Ticket.TYPE_CHOICES)
        status_labels = dict(Ticket.STATUS_CHOICES)
        priority_labels = dict(Ticket.PRIORITY_CHOICES)

        for row, ticket in enumerate(qs, 2):
            col = 1
            ws.cell(row=row, column=col, value=ticket.pk); col += 1
            ws.cell(row=row, column=col, value=ticket.title); col += 1
            ws.cell(row=row, column=col, value=str(type_labels.get(ticket.type, ticket.type))); col += 1
            ws.cell(row=row, column=col, value=str(status_labels.get(ticket.status, ticket.status))); col += 1
            ws.cell(row=row, column=col, value=str(priority_labels.get(ticket.priority, ticket.priority))); col += 1
            ws.cell(row=row, column=col, value=ticket.area.name if ticket.area else ''); col += 1
            ws.cell(row=row, column=col, value=ticket.company.name if ticket.company else ''); col += 1
            ws.cell(row=row, column=col, value=str(ticket.requester) if ticket.requester else ''); col += 1
            if show_resolver:
                ws.cell(row=row, column=col, value=str(ticket.resolver) if ticket.resolver else ''); col += 1
            if show_hours:
                ws.cell(row=row, column=col, value=float(ticket.hours_sum) if ticket.hours_sum else 0); col += 1
            created = localtime(ticket.created_at)
            ws.cell(row=row, column=col, value=created.replace(tzinfo=None))
            ws.cell(row=row, column=col).number_format = 'DD.MM.YYYY'

        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="tikety.xlsx"'
        return response


class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = 'tickets/ticket_detail.html'

    def get_object(self):
        ticket = get_object_or_404(Ticket, pk=self.kwargs['pk'])
        user = self.request.user
        # Kontrola přístupu
        if user.has_role(UserRole.ADMIN):
            return ticket
        if user.has_role(UserRole.MANAGER):
            if user.can_see_ticket_as_manager(ticket):
                return ticket
            messages.error(self.request, _('Nemáte přístup k tomuto tiketu.'))
            raise PermissionError()
        if user.has_role(UserRole.RESOLVER) and (ticket.resolver == user or (ticket.status == Ticket.STATUS_NEW and user.can_handle_ticket_area(ticket))):
            return ticket
        if user.has_role(UserRole.SALES) and ticket.sales == user:
            return ticket
        if user.has_role(UserRole.REQUESTER) and ticket.requester == user:
            return ticket
        messages.error(self.request, _('Nemáte přístup k tomuto tiketu.'))
        raise PermissionError()

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except Http404:
            messages.error(request, _('Tiket neexistuje.'))
            return redirect('tickets:list')
        except PermissionError:
            return redirect('tickets:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ticket = self.object
        user = self.request.user
        ctx['comment_form'] = CommentForm()
        ctx['timelog_form'] = TimeLogForm()
        ctx['comments'] = ticket.comments.select_related('author').all()
        ctx['time_logs'] = ticket.time_logs.select_related('user').all()
        ctx['resolve_form'] = ResolveForm(instance=ticket)
        ctx['reject_form'] = RejectForm(instance=ticket)
        ctx['assign_resolver_form'] = AssignResolverForm(instance=ticket)
        ctx['assign_sales_form'] = AssignSalesForm(instance=ticket)
        ctx['change_type_form'] = ChangeTypeForm(instance=ticket)
        ctx['can_edit'] = _can_edit_ticket(user, ticket)
        if user.has_role(UserRole.ADMIN):
            ctx['can_comment'] = True
        elif user.has_role(UserRole.MANAGER):
            ctx['can_comment'] = user.can_see_ticket_as_manager(ticket)
        elif user.has_role(UserRole.RESOLVER):
            ctx['can_comment'] = ticket.resolver == user
        elif user.has_role(UserRole.SALES):
            ctx['can_comment'] = ticket.sales == user
        elif user.has_role(UserRole.REQUESTER):
            ctx['can_comment'] = ticket.requester == user
        else:
            ctx['can_comment'] = False
        show_hours = user.has_role(UserRole.MANAGER, UserRole.RESOLVER, UserRole.SALES, UserRole.ADMIN)
        ctx['show_hours'] = show_hours
        ctx['has_timelogs'] = ticket.time_logs.exists()
        # Historie změn — interní pole skryta před žadatelem
        history_qs = ticket.history.select_related('user').all()
        if not show_hours:
            history_qs = history_qs.exclude(field__in=TicketChange.INTERNAL_FIELDS)
        ctx['ticket_history'] = history_qs
        prev_pk, next_pk = _get_adjacent_tickets(user, ticket)
        ctx['prev_ticket_pk'] = prev_pk
        ctx['next_ticket_pk'] = next_pk
        # Přílohy
        attachments = ticket.attachments.select_related('uploaded_by').all()
        ctx['attachments'] = attachments
        ctx['can_add_attachment'] = _can_add_attachment(user, ticket)
        ctx['attachment_form'] = AttachmentUploadForm()
        ctx['deletable_attachment_pks'] = {
            att.pk for att in attachments if _can_delete_attachment(user, att)
        }
        return ctx


class TicketCreateView(LoginRequiredMixin, CreateView):
    model = Ticket
    form_class = TicketCreateForm
    template_name = 'tickets/ticket_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        ticket = form.save(commit=False)
        ticket.requester = self.request.user
        user = self.request.user

        # Správce/admin si vybírají firmu ve formuláři, žadatel má svou
        if user.has_role(UserRole.MANAGER, UserRole.ADMIN):
            ticket.company = form.cleaned_data.get('company')
        else:
            ticket.company = user.company

        # Bezpečná kontrola bez přístupu přes FK descriptor
        if not ticket.company_id:
            messages.error(self.request, _('Váš účet nemá přiřazenou firmu. Kontaktujte administrátora.'))
            return self.form_invalid(form)

        ticket.save()
        _log_change(ticket, self.request.user, TicketChange.FIELD_CREATED)

        # Zpracování příloh (volitelné, multiple file input)
        for f in self.request.FILES.getlist('files'):
            err = _validate_upload(f)
            if err:
                messages.warning(self.request, err)
                continue
            TicketAttachment.objects.create(
                ticket=ticket,
                file=f,
                original_name=f.name,
                uploaded_by=self.request.user,
            )
            _log_change(ticket, self.request.user, TicketChange.FIELD_ATTACHMENT_ADDED,
                        new_value=f.name)

        from apps.notifications.tasks import notify_new_ticket
        notify_new_ticket.delay(ticket.pk)
        messages.success(self.request, _('Tiket byl vytvořen.'))
        return redirect('tickets:detail', pk=ticket.pk)


class TicketUpdateView(LoginRequiredMixin, UpdateView):
    model = Ticket
    form_class = TicketUpdateForm
    template_name = 'tickets/ticket_form.html'

    def dispatch(self, request, *args, **kwargs):
        ticket = get_object_or_404(Ticket, pk=kwargs['pk'])
        if not _can_edit_ticket(request.user, ticket):
            messages.error(request, _('Tiket nelze upravit v aktuálním stavu.'))
            return redirect('tickets:detail', pk=ticket.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['area'] = self.object.area
        return kwargs

    def form_valid(self, form):
        old = Ticket.objects.select_related('area').get(pk=self.object.pk)
        old_description = old.description
        old_title = old.title

        ticket = form.save(commit=False)

        if old.type != ticket.type:
            ticket.save()
            ticket.change_type(ticket.type)
        else:
            ticket.save()

        user = self.request.user
        if old.type != ticket.type:
            _log_change(ticket, user, TicketChange.FIELD_TYPE,
                        _type_label(old.type), _type_label(ticket.type))
        if old.priority != ticket.priority:
            _log_change(ticket, user, TicketChange.FIELD_PRIORITY,
                        _priority_label(old.priority), _priority_label(ticket.priority))
        if old.area_id != ticket.area_id:
            _log_change(ticket, user, TicketChange.FIELD_AREA,
                        str(old.area) if old.area else '—',
                        str(ticket.area) if ticket.area else '—')
        if old_title != ticket.title:
            _log_change(ticket, user, TicketChange.FIELD_TITLE,
                        old_title[:200], ticket.title[:200])
        if old_description != ticket.description:
            _log_change(ticket, user, TicketChange.FIELD_DESCRIPTION,
                        old_value=old_description)
        if old.work_category_id != ticket.work_category_id:
            _log_change(ticket, user, TicketChange.FIELD_WORK_CATEGORY,
                        str(old.work_category) if old.work_category else '',
                        str(ticket.work_category) if ticket.work_category else '')

        messages.success(self.request, _('Tiket byl uložen.'))
        return redirect('tickets:detail', pk=ticket.pk)


# ------------------------------------------------------------------ #
# HTMX akce                                                           #
# ------------------------------------------------------------------ #

class AssignResolverView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not _manager_has_ticket_access(request.user, ticket):
            return HttpResponse(_('Nedostatečná oprávnění.'), status=403)
        old_resolver = ticket.resolver
        old_status = ticket.status
        form = AssignResolverForm(request.POST, instance=ticket)
        if form.is_valid():
            ticket = form.save(commit=False)
            if ticket.status in (Ticket.STATUS_NEW, Ticket.STATUS_OFFER):
                ticket.to_in_progress()
                ticket.save()
                from apps.notifications.tasks import notify_status_change, notify_assigned_to_resolver
                notify_status_change.delay(ticket.pk)
                notify_assigned_to_resolver.delay(ticket.pk)
            elif ticket.status == Ticket.STATUS_IN_PROGRESS:
                ticket.save()
                from apps.notifications.tasks import notify_assigned_to_resolver
                notify_assigned_to_resolver.delay(ticket.pk)
            _log_change(ticket, request.user, TicketChange.FIELD_RESOLVER,
                        str(old_resolver) if old_resolver else '', str(ticket.resolver) if ticket.resolver else '')
            if old_status != ticket.status:
                _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                            _status_label(old_status), _status_label(ticket.status))
            messages.success(request, _('Řešitel byl přiřazen.'))
        return redirect('tickets:detail', pk=pk)


class AssignSalesView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not _manager_has_ticket_access(request.user, ticket):
            return HttpResponse(_('Nedostatečná oprávnění.'), status=403)
        if ticket.type != Ticket.TYPE_DEVELOPMENT:
            messages.error(request, _('Obchodníka lze přiřadit pouze k typu „Požadavek na vývoj".'))
            return redirect('tickets:detail', pk=pk)
        old_sales = ticket.sales
        old_status = ticket.status
        form = AssignSalesForm(request.POST, instance=ticket)
        if form.is_valid():
            ticket = form.save(commit=False)
            if ticket.status == Ticket.STATUS_NEW:
                ticket.to_offer_prep()
                ticket.save()
                from apps.notifications.tasks import notify_status_change, notify_assigned_to_sales
                notify_status_change.delay(ticket.pk)
                notify_assigned_to_sales.delay(ticket.pk)
            elif ticket.status == Ticket.STATUS_OFFER:
                ticket.save()
                from apps.notifications.tasks import notify_assigned_to_sales
                notify_assigned_to_sales.delay(ticket.pk)
            _log_change(ticket, request.user, TicketChange.FIELD_SALES,
                        str(old_sales) if old_sales else '', str(ticket.sales) if ticket.sales else '')
            if old_status != ticket.status:
                _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                            _status_label(old_status), _status_label(ticket.status))
            messages.success(request, _('Obchodník byl přiřazen.'))
        return redirect('tickets:detail', pk=pk)


class TakeTicketView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not request.user.has_role(UserRole.RESOLVER):
            messages.error(request, _('Nedostatečná oprávnění.'))
            return redirect('tickets:detail', pk=pk)
        if ticket.status != Ticket.STATUS_NEW:
            messages.error(request, _('Tiket již nelze převzít.'))
            return redirect('tickets:detail', pk=pk)
        if not request.user.can_handle_ticket_area(ticket):
            messages.error(request, _('Nemáte oprávnění převzít tiket z této oblasti.'))
            return redirect('tickets:detail', pk=pk)
        old_status = ticket.status
        ticket.resolver = request.user
        ticket.to_in_progress()
        ticket.save()
        from apps.notifications.tasks import notify_status_change, notify_assigned_to_resolver
        notify_status_change.delay(ticket.pk)
        notify_assigned_to_resolver.delay(ticket.pk)
        _log_change(ticket, request.user, TicketChange.FIELD_RESOLVER,
                    '', str(request.user))
        _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                    _status_label(old_status), _status_label(ticket.status))
        messages.success(request, _('Tiket byl převzat.'))
        return redirect('tickets:detail', pk=pk)


class ResolveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        can_resolve = (
            _manager_has_ticket_access(request.user, ticket) or
            (request.user.has_role(UserRole.RESOLVER) and ticket.resolver == request.user)
        )
        if not can_resolve:
            messages.error(request, _('Nedostatečná oprávnění.'))
            return redirect('tickets:detail', pk=pk)
        form = ResolveForm(request.POST, instance=ticket)
        if form.is_valid():
            hours_str = request.POST.get('hours', '').strip()
            if not hours_str and not ticket.time_logs.exists():
                messages.error(request, _('Zadejte počet odpracovaných hodin.'))
                return redirect('tickets:detail', pk=pk)
            old_status = ticket.status
            ticket = form.save(commit=False)
            ticket.to_resolved()
            ticket.save()
            _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                        _status_label(old_status), _status_label(ticket.status))
            if hours_str:
                try:
                    hours = float(hours_str)
                    if hours > 0:
                        timelog_user = ticket.resolver if ticket.resolver_id else request.user
                        TimeLog.objects.create(ticket=ticket, user=timelog_user, hours=hours)
                except ValueError:
                    pass
            from apps.notifications.tasks import notify_ticket_closed
            notify_ticket_closed.delay(ticket.pk, closed_as='resolved')
            messages.success(request, _('Tiket byl vyřešen.'))
        else:
            messages.error(request, _('Vyplňte způsob vyřešení.'))
        return redirect('tickets:detail', pk=pk)


class RejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not _manager_has_ticket_access(request.user, ticket):
            messages.error(request, _('Nedostatečná oprávnění.'))
            return redirect('tickets:detail', pk=pk)
        old_status = ticket.status
        form = RejectForm(request.POST, instance=ticket)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.to_rejected()
            ticket.save()
            _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                        _status_label(old_status), _status_label(ticket.status))
            from apps.notifications.tasks import notify_ticket_closed
            notify_ticket_closed.delay(ticket.pk, closed_as='rejected')
            messages.success(request, _('Tiket byl zamítnut.'))
        else:
            messages.error(request, _('Vyplňte důvod zamítnutí.'))
        return redirect('tickets:detail', pk=pk)


class ReopenView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not _manager_has_ticket_access(request.user, ticket):
            messages.error(request, _('Nedostatečná oprávnění.'))
            return redirect('tickets:detail', pk=pk)
        old_status = ticket.status
        target = request.POST.get('target', 'in_progress')
        if target == 'offer_prep' and ticket.type == Ticket.TYPE_DEVELOPMENT:
            ticket.reopen_to_offer_prep()
        else:
            ticket.reopen_to_in_progress()
        ticket.save()
        _log_change(ticket, request.user, TicketChange.FIELD_STATUS,
                    _status_label(old_status), _status_label(ticket.status))
        messages.success(request, _('Tiket byl znovuotevřen.'))
        return redirect('tickets:detail', pk=pk)


class ChangeTypeView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        can_change_type = (
            _manager_has_ticket_access(request.user, ticket) or
            (request.user.has_role(UserRole.RESOLVER) and ticket.resolver == request.user)
        )
        if not can_change_type:
            messages.error(request, _('Nedostatečná oprávnění.'))
            return redirect('tickets:detail', pk=pk)
        form = ChangeTypeForm(request.POST, instance=ticket)
        if form.is_valid():
            old_type = ticket.type
            ticket.change_type(form.cleaned_data['type'])
            _log_change(ticket, request.user, TicketChange.FIELD_TYPE,
                        _type_label(old_type), _type_label(ticket.type))
            messages.success(request, _('Typ tiketu byl změněn.'))
        return redirect('tickets:detail', pk=pk)


class AddCommentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        user = request.user
        # Každá role smí komentovat pouze tikety, ke kterým má přístup
        if user.has_role(UserRole.ADMIN):
            can_comment = True
        elif user.has_role(UserRole.MANAGER):
            can_comment = user.can_see_ticket_as_manager(ticket)
        elif user.has_role(UserRole.RESOLVER):
            can_comment = ticket.resolver == user
        elif user.has_role(UserRole.SALES):
            can_comment = ticket.sales == user
        elif user.has_role(UserRole.REQUESTER):
            can_comment = ticket.requester == user
        else:
            can_comment = False
        if not can_comment:
            messages.error(request, _('Nemáte oprávnění komentovat tento tiket.'))
            return redirect('tickets:list')
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.ticket = ticket
            comment.author = request.user
            comment.save()
            from apps.notifications.tasks import notify_new_comment
            notify_new_comment.delay(comment.pk)
            if request.htmx:
                return render(request, 'tickets/partials/comment_list.html', {
                    'ticket': ticket,
                    'comments': ticket.comments.select_related('author').all(),
                    'comment_form': CommentForm(),
                    'can_comment': can_comment,
                })
        return redirect('tickets:detail', pk=pk)


class AddAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        if not _can_add_attachment(request.user, ticket):
            messages.error(request, _('Nemáte oprávnění přidávat přílohy.'))
            return redirect('tickets:detail', pk=pk)

        files = request.FILES.getlist('files')
        if not files:
            messages.error(request, _('Nevybrali jste žádné soubory.'))
            return redirect('tickets:detail', pk=pk)

        for f in files:
            err = _validate_upload(f)
            if err:
                messages.warning(request, err)
                continue
            TicketAttachment.objects.create(
                ticket=ticket,
                file=f,
                original_name=f.name,
                uploaded_by=request.user,
            )
            _log_change(ticket, request.user, TicketChange.FIELD_ATTACHMENT_ADDED,
                        new_value=f.name)

        if request.htmx:
            attachments = ticket.attachments.select_related('uploaded_by').all()
            return render(request, 'tickets/partials/attachment_list.html', {
                'ticket': ticket,
                'attachments': attachments,
                'can_add_attachment': _can_add_attachment(request.user, ticket),
                'attachment_form': AttachmentUploadForm(),
                'deletable_attachment_pks': {
                    att.pk for att in attachments if _can_delete_attachment(request.user, att)
                },
            })
        return redirect('tickets:detail', pk=pk)


class DeleteAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk, att_pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        attachment = get_object_or_404(TicketAttachment, pk=att_pk, ticket=ticket)
        if not _can_delete_attachment(request.user, attachment):
            messages.error(request, _('Nemáte oprávnění smazat tuto přílohu.'))
            return redirect('tickets:detail', pk=pk)

        original_name = attachment.original_name
        # Smazat fyzický soubor
        try:
            if attachment.file and os.path.isfile(attachment.file.path):
                os.remove(attachment.file.path)
        except (ValueError, OSError):
            pass
        attachment.delete()
        _log_change(ticket, request.user, TicketChange.FIELD_ATTACHMENT_DELETED,
                    new_value=original_name)

        if request.htmx:
            attachments = ticket.attachments.select_related('uploaded_by').all()
            return render(request, 'tickets/partials/attachment_list.html', {
                'ticket': ticket,
                'attachments': attachments,
                'can_add_attachment': _can_add_attachment(request.user, ticket),
                'attachment_form': AttachmentUploadForm(),
                'deletable_attachment_pks': {
                    att.pk for att in attachments if _can_delete_attachment(request.user, att)
                },
            })
        return redirect('tickets:detail', pk=pk)


class DownloadAttachmentView(LoginRequiredMixin, View):
    def get(self, request, pk, att_pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        attachment = get_object_or_404(TicketAttachment, pk=att_pk, ticket=ticket)

        # Kontrola přístupu — stejná logika jako detail tiketu
        user = request.user
        allowed = False
        if user.has_role(UserRole.ADMIN):
            allowed = True
        elif user.has_role(UserRole.MANAGER):
            allowed = user.can_see_ticket_as_manager(ticket)
        elif user.has_role(UserRole.RESOLVER):
            allowed = ticket.resolver == user or (ticket.status == Ticket.STATUS_NEW and user.can_handle_ticket_area(ticket))
        elif user.has_role(UserRole.SALES):
            allowed = ticket.sales == user
        elif user.has_role(UserRole.REQUESTER):
            allowed = ticket.requester == user

        if not allowed:
            messages.error(request, _('Nemáte přístup k této příloze.'))
            return redirect('tickets:list')

        try:
            response = FileResponse(
                open(attachment.file.path, 'rb'),
                as_attachment=True,
                filename=attachment.original_name,
            )
            return response
        except (FileNotFoundError, OSError):
            messages.error(request, _('Soubor nebyl nalezen.'))
            return redirect('tickets:detail', pk=pk)


class AddTimeLogView(LoginRequiredMixin, View):
    def post(self, request, pk):
        ticket = get_object_or_404(Ticket, pk=pk)
        user = request.user
        # Správce smí zapisovat čas jen pokud je zároveň přiřazen jako řešitel nebo obchodník
        can_log = (
            (user.has_role(UserRole.MANAGER) and (ticket.resolver == user or ticket.sales == user)) or
            (user.has_role(UserRole.RESOLVER) and ticket.resolver == user) or
            (user.has_role(UserRole.SALES) and ticket.sales == user)
        )
        if not can_log:
            messages.error(request, _('Nemáte oprávnění zapisovat čas.'))
            return redirect('tickets:detail', pk=pk)
        form = TimeLogForm(request.POST)
        if form.is_valid():
            log = form.save(commit=False)
            log.ticket = ticket
            log.user = request.user
            log.save()
            messages.success(request, _('Čas byl zapsán.'))
            if request.htmx:
                return render(request, 'tickets/partials/timelog_list.html', {
                    'ticket': ticket,
                    'time_logs': ticket.time_logs.select_related('user').all(),
                })
        return redirect('tickets:detail', pk=pk)


# ------------------------------------------------------------------ #
# Správa kategorií práce (Administrátor)                              #
# ------------------------------------------------------------------ #

def _admin_only(request):
    if not request.user.is_authenticated or not request.user.has_role(UserRole.ADMIN):
        messages.error(request, _('Nemáte oprávnění.'))
        return True
    return False


class WorkCategoryListView(LoginRequiredMixin, ListView):
    model = WorkCategory
    template_name = 'tickets/work_category_list.html'
    context_object_name = 'categories'

    def dispatch(self, request, *args, **kwargs):
        if _admin_only(request):
            return redirect('tickets:list')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return WorkCategory.objects.prefetch_related('areas').all()


class WorkCategoryCreateView(LoginRequiredMixin, CreateView):
    model = WorkCategory
    form_class = WorkCategoryAdminForm
    template_name = 'tickets/work_category_form.html'
    success_url = reverse_lazy('tickets:work_category_list')

    def dispatch(self, request, *args, **kwargs):
        if _admin_only(request):
            return redirect('tickets:list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, _('Kategorie byla vytvořena.'))
        return super().form_valid(form)


class WorkCategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = WorkCategory
    form_class = WorkCategoryAdminForm
    template_name = 'tickets/work_category_form.html'
    success_url = reverse_lazy('tickets:work_category_list')

    def dispatch(self, request, *args, **kwargs):
        if _admin_only(request):
            return redirect('tickets:list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, _('Kategorie byla uložena.'))
        return super().form_valid(form)


